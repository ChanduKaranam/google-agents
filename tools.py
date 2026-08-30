"""The batch ledger.

Lodestar is asked to keep a high-level overview of a batch across many turns —
how many leads came in, how many were analysed, how many were reached, which
ones are owed a retry. A model asked to carry that in its head reports numbers
that drift, and the numbers here are the ones a human is told before a hundred
strangers get phoned. So the ledger is real: these tools write it to session
state and read it back, and the orchestrator is told never to quote a figure it
did not get from `batch_status`.

Every tool returns a plain dict with a `status` of 'success' or 'error'.
"""

import csv
import io
import json
import logging
import re
import uuid
from datetime import datetime, timezone

from google.adk.tools.tool_context import ToolContext

from . import config

logger = logging.getLogger(__name__)

# Where a lead has got to. A lead moves forward only through these tools.
RECEIVED = 'received'    # in the batch, not yet analysed
ANALYSED = 'analysed'    # has a profile and a recommended policy
CALLED = 'called'        # reached — the outreach specialist got through
IN_PROGRESS = 'in_progress'  # dialled, still live; result not known yet
UNATTEMPTED = 'unattempted'  # the call did not happen; retryable
FAILED = 'failed'        # the call was attempted and did not connect
FLAGGED = 'flagged'      # out of the pipeline, waiting on a human


def _field(row: dict, *names: str):
    """One field from a spreadsheet row, however the column was capitalised.

    Lead sheets arrive with "Name", "name", "Phone Number", "phone_number" and
    worse in the same column across two files from the same agency. Matching
    exactly meant a header case change silently blanked every phone number.
    """
    normalised = {
        str(key).strip().lower().replace(' ', '_').replace('-', '_'): value
        for key, value in row.items()
    }
    for name in names:
        value = normalised.get(name)
        if value not in (None, ''):
            return value
    return None


# --- deterministic underwriting ---------------------------------------------
#
# Gap, priority and product are computed here, not by the model, so the same
# sheet always produces the same numbers. The analysis agent's job is the
# "Why" sentence and the pitch notes — never the arithmetic.

HOT, WARM, COLD = 'HOT', 'WARM', 'COLD'


def _clean_key(key: str) -> str:
    return re.sub(r'[^a-z]', '', str(key).lower())


def _row_value(row: dict, *fragments: str):
    """A field matched by substring of its cleaned header.

    "Income (S$K)", "Annual Income (SGD)" and "income_sgd" all reach the same
    fragment 'income'. First fragment that matches any key wins.
    """
    cleaned = {_clean_key(k): v for k, v in row.items()}
    for fragment in fragments:
        for key, value in cleaned.items():
            if fragment in key and value not in (None, ''):
                return value
    return None


def _num(value) -> float:
    match = re.search(r'-?\d+(?:\.\d+)?', str(value or '').replace(',', ''))
    return float(match.group()) if match else 0.0


def _thousands(value: float) -> float:
    # "120" in an S$K column and "120000" in an SGD column mean the same money.
    return value / 1000 if value >= 10000 else value


def _assess(row: dict) -> dict:
    """Gap (S$K), priority and product for one raw lead row."""
    income_k = _thousands(_num(_row_value(row, 'income')))
    cover_k = _thousands(_num(_row_value(row, 'cover')))
    age = int(_num(_row_value(row, 'age')))
    dependents = int(_num(_row_value(row, 'dependent', 'dep')))
    smoker = str(_row_value(row, 'tobacco', 'smok') or '').strip().lower() in (
        'y', 'yes', 'true', 'smoker', '1')
    event = str(_row_value(row, 'event') or '').strip().lower()

    gap_k = max(0.0, round(config.INCOME_MULTIPLE * income_k - cover_k))

    qualifying_event = any(w in event for w in ('child', 'loan', 'marriage'))
    if (qualifying_event and gap_k > config.HOT_MIN_GAP_K
            and cover_k < config.HOT_MAX_COVER_RATIO * income_k):
        priority = HOT
    elif gap_k > config.WARM_MIN_GAP_K and (dependents >= 1 or smoker):
        priority = WARM
    else:
        priority = COLD

    if 'child' in event:
        policy = 'Term + Child Education Plan'
    elif 'loan' in event:
        policy = 'Mortgage-linked Term'
    elif 'marriage' in event:
        policy = 'Term + Whole Life'
    elif age >= config.LEGACY_MIN_AGE and cover_k >= config.LEGACY_MIN_COVER_K:
        policy = 'Retirement / Legacy Plan'
    elif smoker:
        policy = 'Term top-up + CI rider'
    elif age < config.STARTER_MAX_AGE and dependents == 0:
        policy = 'Savings + CI starter'
    elif gap_k <= config.SMALL_GAP_K:
        policy = 'Health / CI top-up'
    else:
        policy = 'Term top-up'
    if smoker and 'CI' not in policy:
        policy += ' + CI rider'

    # A HOT term lead gets a suggested sum assured: the gap, rounded up to the
    # nearest S$100K, so "gap 980" is pitched as "Term S$1M".
    if priority == HOT and policy.startswith('Term'):
        sum_k = int(-(-gap_k // 100) * 100)
        label = f'S${sum_k / 1000:g}M' if sum_k >= 1000 else f'S${sum_k:d}K'
        policy = policy.replace('Term', f'Term {label}', 1)

    return {
        'income_k': income_k,
        'cover_k': cover_k,
        'age': age,
        'dependents': dependents,
        'smoker': smoker,
        'life_event': event,
        'gap_k': gap_k,
        'priority': priority,
        'policy': policy,
    }


def _parse_rows(text: str) -> list[dict]:
    """Rows out of whatever delimited text the model passed through.

    The lead sheet reaches this tool as the text of the uploaded file rather
    than as a parsed list of dicts, and deliberately so: a list-of-dicts
    argument makes Gemini emit its function call as a code block, which arrives
    as MALFORMED_FUNCTION_CALL and loses the entire batch. One string argument
    it can always serialise.

    Comma, pipe, tab and semicolon are all accepted, because a file pasted into
    a chat window has usually stopped being a CSV by the time it arrives.
    """
    text = (text or '').strip()
    if not text:
        return []
    # A fenced block, if the model wrapped it in one.
    fenced = re.match(r'^```[a-zA-Z]*\s*\n(.*?)\n?```$', text, re.S)
    if fenced:
        text = fenced.group(1).strip()

    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        return []

    header = lines[0]
    delimiter = max(',|\t;', key=header.count)
    if header.count(delimiter) == 0:
        return []

    rows = list(csv.DictReader(io.StringIO('\n'.join(lines)), delimiter=delimiter))
    cleaned = []
    for row in rows:
        # DictReader gives None keys for surplus columns and None values for
        # short rows; neither should reach the ledger.
        item = {
            str(k).strip(): ('' if v is None else str(v).strip())
            for k, v in row.items() if k is not None
        }
        if any(item.values()):
            cleaned.append(item)
    return cleaned


def _xlsx_rows(data: bytes) -> list[dict] | None:
    """Rows out of an .xlsx, or None if the bytes are not one.

    The first sheet's first row is the header. Hidden columns are read like
    any other — that is where the demo sheet keeps its phone numbers.
    """
    if not data[:2] == b'PK':
        return None
    try:
        from openpyxl import load_workbook
        sheet = load_workbook(io.BytesIO(data), read_only=True,
                              data_only=True).worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else '' for h in next(rows)]
    except Exception:  # noqa: BLE001 - not an xlsx after all
        return None
    parsed = []
    for row in rows:
        item = {
            h: ('' if v is None else str(v).strip())
            for h, v in zip(header, row) if h
        }
        if any(item.values()):
            parsed.append(item)
    return parsed


def _artifact_bytes(part) -> bytes | None:
    data = getattr(getattr(part, 'inline_data', None), 'data', None)
    if isinstance(data, (bytes, bytearray)):
        return bytes(data)
    return None


def _artifact_text(part) -> str | None:
    """The text of an attached file, or None if it is not text at all.

    An attachment arrives as a `types.Part`: usually `inline_data` holding raw
    bytes, occasionally already decoded into `text`. A CSV saved out of Excel
    is frequently cp1252 rather than UTF-8, and is often carrying a BOM, so
    both are tried before giving up.
    """
    text = getattr(part, 'text', None)
    if text:
        return text

    inline = getattr(part, 'inline_data', None)
    data = getattr(inline, 'data', None)
    if not data:
        return None
    if isinstance(data, str):
        return data

    for encoding in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
        try:
            return bytes(data).decode(encoding)
        except (UnicodeDecodeError, ValueError):
            continue
    return None


def _now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec='seconds')


def _batch(tool_context: ToolContext) -> dict | None:
    return tool_context.state.get(config.BATCH_STATE_KEY)


def _save(tool_context: ToolContext, batch: dict) -> None:
    tool_context.state[config.BATCH_STATE_KEY] = batch


def _no_batch() -> dict:
    return {
        'status': 'error',
        'error_message': (
            'No batch is open. Call open_batch with the raw leads before '
            'anything else.'
        ),
    }


def _counts(batch: dict) -> dict:
    """The tally the orchestrator is allowed to quote."""
    leads = batch['leads']
    stages = [lead['stage'] for lead in leads.values()]
    return {
        'total': len(leads),
        'hot': sum(1 for l in leads.values() if l.get('priority') == HOT),
        'warm': sum(1 for l in leads.values() if l.get('priority') == WARM),
        'cold': sum(1 for l in leads.values() if l.get('priority') == COLD),
        'combined_gap_sgd_k': round(
            sum(l.get('gap_k', 0) for l in leads.values())),
        'analysed': sum(1 for s in stages if s != RECEIVED and s != FLAGGED),
        'ready_to_call': sum(
            1 for lead in leads.values()
            if lead['stage'] == ANALYSED and lead['attempts'] == 0
        ),
        'contacted': sum(1 for s in stages if s == CALLED),
        'awaiting_result': sum(1 for s in stages if s == IN_PROGRESS),
        'simulated_calls': sum(1 for lead in leads.values() if lead['mock']),
        'pending_retry': sum(
            1
            for lead in leads.values()
            if lead['stage'] in (UNATTEMPTED, FAILED)
            and lead['attempts'] < config.MAX_CALL_ATTEMPTS
            and lead['retry_rounds'] < config.MAX_CALL_ATTEMPTS
        ),
        'flagged': sum(1 for s in stages if s == FLAGGED),
        'awaiting_analysis': sum(1 for s in stages if s == RECEIVED),
    }


_UNREADABLE = (
    'No leads could be read from that. Expected a header row naming the '
    'columns and one row per lead, separated by commas, pipes or tabs.'
)


def _open_from_rows(leads: list[dict], source: str,
                    tool_context: ToolContext) -> dict:
    """Start a ledger for rows that have already been parsed.

    Shared by both ways in — a sheet pasted into the conversation and a file
    the user attached — so the two cannot drift apart.
    """

    batch = {
        'batch_id': f'batch-{uuid.uuid4().hex[:8]}',
        'source': source,
        'opened_at': _now(),
        'confirmed': not config.CONFIRM_BEFORE_CALLS,
        'leads': {},
    }

    for index, row in enumerate(leads, start=1):
        row = row if isinstance(row, dict) else {'value': row}
        # Identity is best-effort and never invented. A row with no usable name
        # or phone still gets a ledger entry — it has to, or the totals stop
        # matching the file — and analysis is what decides whether it is
        # workable.
        lead_id = str(
            _field(row, 'lead_id', 'id') or f'L{index:03d}'
        )
        assessment = _assess(row)
        batch['leads'][lead_id] = {
            'lead_id': lead_id,
            'name': str(_field(row, 'name', 'full_name', 'lead_name') or ''),
            'phone': str(_field(
                row, 'phone', 'phone_number', 'mobile', 'contact',
                'contact_number', 'mobile_number',
            ) or ''),
            'stage': RECEIVED,
            'gap_k': assessment['gap_k'],
            'priority': assessment['priority'],
            'age': assessment['age'],
            'policy': assessment['policy'],
            'reasoning': '',
            'pitch_notes': '',
            'profile': {},
            'attempts': 0,
            'retry_rounds': 0,
            'call_id': '',
            'mock': False,
            'last_status': '',
            'note': '',
            'raw': row,
        }

    _save(tool_context, batch)
    no_phone = [l['lead_id'] for l in batch['leads'].values() if not l['phone']]
    logger.info('opened %s from %r with %d leads (%d without a number)',
                batch['batch_id'], source, len(batch['leads']), len(no_phone))
    return {
        'status': 'success',
        'batch_id': batch['batch_id'],
        'source': source,
        'counts': _counts(batch),
        'confirmation_required': config.CONFIRM_BEFORE_CALLS,
        'chunk_size': config.ANALYSIS_CHUNK_SIZE,
        # Said here rather than discovered at dialling time. A whole file with
        # no phone column is the likeliest version of this, and it is worth
        # knowing before the analysis runs.
        'leads_without_a_phone_number': no_phone,
        'calls_are_simulated': config.mock_calls(),
    }


def open_batch(leads_csv: str, source: str,
               tool_context: ToolContext) -> dict:
    """Register a batch of raw leads pasted into the conversation as text.

    Use this when the lead rows are in the message itself. When the user
    attached a file instead, use `open_batch_from_file` — it reads the file
    directly and does not need the rows repeated.

    Args:
        leads_csv: The lead sheet as delimited text, exactly as it arrived —
            a header row naming the columns, then one row per lead. Comma,
            pipe, tab and semicolon all work.
        source: Where the batch came from, e.g. the file name.

    Returns:
        The batch id and the opening counts.
    """
    leads = _parse_rows(leads_csv)
    if not leads:
        return {'status': 'error', 'error_message': _UNREADABLE}
    return _open_from_rows(leads, source, tool_context)


async def list_uploaded_files(tool_context: ToolContext) -> dict:
    """List the files the user has attached to this conversation.

    Call this first when someone says they have uploaded a lead sheet. It
    returns the filenames; pass the one you want to `open_batch_from_file`.

    Returns:
        The filenames attached to this session, newest last.
    """
    try:
        names = await tool_context.list_artifacts()
    except ValueError:
        # No artifact service wired up on this deployment.
        return {
            'status': 'error',
            'error_message': (
                'This deployment cannot see attached files. Ask the user to '
                'paste the lead rows into the message instead.'
            ),
        }
    except Exception as exc:  # noqa: BLE001 - surfaced to the model, not raised
        logger.exception('listing artifacts failed')
        return {'status': 'error', 'error_message': f'Could not list the attached files: {exc}'}

    return {
        'status': 'success',
        'files': names,
        'count': len(names),
    }


async def open_batch_from_file(filename: str,
                               tool_context: ToolContext) -> dict:
    """Register a batch of raw leads from a file the user attached.

    Reads the file itself. The lead rows never pass back through the
    conversation, which is the point: a sheet of any size opens the same way,
    and nothing is retyped or summarised on the way in.

    Args:
        filename: The attached file to read, as `list_uploaded_files` gave it.
            If there is exactly one attached file, "" reads that one.

    Returns:
        The batch id and the opening counts.
    """
    try:
        name = (filename or '').strip()
        if not name:
            names = await tool_context.list_artifacts()
            if len(names) != 1:
                return {
                    'status': 'error',
                    'error_message': (
                        'Say which file to read. Attached: '
                        + (', '.join(names) if names else 'none')
                    ),
                }
            name = names[0]

        part = await tool_context.load_artifact(name)
    except ValueError:
        return {
            'status': 'error',
            'error_message': (
                'This deployment cannot see attached files. Ask the user to '
                'paste the lead rows into the message instead.'
            ),
        }
    except Exception as exc:  # noqa: BLE001 - surfaced to the model, not raised
        logger.exception('loading artifact %r failed', filename)
        return {'status': 'error', 'error_message': f'Could not read {filename!r}: {exc}'}

    if part is None:
        return {
            'status': 'error',
            'error_message': f'There is no attached file called {name!r}.',
        }

    data = _artifact_bytes(part)
    if data is not None:
        rows = _xlsx_rows(data)
        if rows is not None:
            if not rows:
                return {'status': 'error', 'error_message': _UNREADABLE}
            return _open_from_rows(rows, name, tool_context)

    text = _artifact_text(part)
    if text is None:
        mime = getattr(getattr(part, 'inline_data', None), 'mime_type', '') or 'unknown'
        return {
            'status': 'error',
            'error_message': (
                f'{name!r} is not something I can read as a lead sheet '
                f'(it looks like {mime}). A CSV or other delimited text file '
                f'is what this expects — an .xlsx has to be exported to CSV '
                f'first.'
            ),
        }

    leads = _parse_rows(text)
    if not leads:
        return {'status': 'error', 'error_message': _UNREADABLE}
    return _open_from_rows(leads, name, tool_context)


def record_analysis(profiles_json: str, tool_context: ToolContext) -> dict:
    """Record what the Policy Analysis Agent returned.

    Anything missing a lead id or a recommended policy is flagged rather than
    recorded, and comes back in `incomplete` — those leads are not called.

    Args:
        profiles_json: The JSON array the analysis specialist returned, passed
            through as text exactly as it came. One object per lead, each with
            a `lead_id` and a `recommended_policy`; `reasoning`,
            `pitch_notes`, `extracted_profile` and `lead_name` are carried
            through to the caller if present.

    Returns:
        The counts after recording, plus any profiles that could not be used.
    """
    batch = _batch(tool_context)
    if batch is None:
        return _no_batch()

    # Taken as text rather than as a list of dicts for the same reason
    # `open_batch` takes the sheet as text: a nested argument this size makes
    # Gemini emit the call as a code block, and the chunk is lost.
    profiles = profiles_json
    if isinstance(profiles, str):
        raw = profiles.strip()
        fenced = re.match(r'^```[a-zA-Z]*\s*\n(.*?)\n?```$', raw, re.S)
        if fenced:
            raw = fenced.group(1).strip()
        try:
            profiles = json.loads(raw) if raw else []
        except json.JSONDecodeError as exc:
            return {
                'status': 'error',
                'error_message': (
                    f'That analysis was not readable JSON ({exc.msg} at line '
                    f'{exc.lineno}). Send the array exactly as the analysis '
                    f'specialist returned it. Nothing was recorded.'
                ),
            }
    if isinstance(profiles, dict):
        profiles = [profiles]
    if not isinstance(profiles, list):
        return {
            'status': 'error',
            'error_message': 'Expected a JSON array of profiles.',
        }

    recorded, incomplete, unknown = [], [], []
    for profile in profiles or []:
        if not isinstance(profile, dict):
            incomplete.append({'profile': profile, 'reason': 'not a record'})
            continue
        lead_id = str(profile.get('lead_id') or '')
        lead = batch['leads'].get(lead_id)
        if lead is None:
            # A lead id that was never in the batch. Recording it would make
            # the totals disagree with the file the batch came from.
            unknown.append(lead_id or '(missing lead_id)')
            continue
        # The ledger's computed policy is authoritative — the analysis agent
        # explains it, it does not choose it. The agent's own answer is only a
        # fallback for a row the deterministic rules could not read.
        policy = lead['policy'] or str(
            profile.get('recommended_policy') or profile.get('policy') or ''
        ).strip()
        if not policy:
            lead['stage'] = FLAGGED
            lead['note'] = (
                str(profile.get('reasoning') or '').strip()
                or 'analysis returned no recommended policy'
            )
            incomplete.append({'lead_id': lead_id, 'reason': lead['note']})
            continue

        lead['stage'] = ANALYSED
        lead['policy'] = policy
        lead['reasoning'] = str(profile.get('reasoning') or '').strip()
        # Written as bullets, and sometimes as a list of them. Both reach the
        # caller as one block of text; this is the only place the difference
        # exists, so it is flattened here rather than in the outreach prompt.
        pitch = profile.get('pitch_notes') or profile.get('note') or ''
        if isinstance(pitch, (list, tuple)):
            pitch = '\n'.join(str(line).strip() for line in pitch if str(line).strip())
        lead['pitch_notes'] = str(pitch).strip()
        lead['profile'] = profile.get('extracted_profile') or {}
        lead['note'] = lead['reasoning']
        # A name from the file is what the caller should use; the analysis
        # agent's `lead_name` fills in only where the row had none.
        if not lead['name']:
            lead['name'] = str(profile.get('lead_name') or '').strip()
        recorded.append(lead_id)

    # Leads the specialist simply did not answer for. Silence is not a pass:
    # left at `received` they would sit in the batch forever, counted as
    # neither reached nor flagged.
    missing = [
        lead_id for lead_id, lead in batch['leads'].items()
        if lead['stage'] == RECEIVED
    ]

    _save(tool_context, batch)
    return {
        'status': 'success',
        'recorded': len(recorded),
        'incomplete': incomplete,
        'unknown_lead_ids': unknown,
        'still_awaiting_analysis': missing,
        'counts': _counts(batch),
    }


def leads_ready_to_call(tool_context: ToolContext) -> dict:
    """The leads the Outreach Agent should be given next.

    Returns analysed leads that have never been called, and separately the ones
    owed a retry. Read this rather than working out the list yourself.
    """
    batch = _batch(tool_context)
    if batch is None:
        return _no_batch()

    if not batch['confirmed']:
        return {
            'status': 'error',
            'error_message': (
                'This batch has not been confirmed for calling yet. Ask the '
                'human to confirm, then call confirm_calling.'
            ),
            'counts': _counts(batch),
        }

    def summary(lead):
        return {
            'lead_id': lead['lead_id'],
            'name': lead['name'],
            'phone': lead['phone'],
            'priority': lead.get('priority', ''),
            'gap_k': lead.get('gap_k', 0),
            'policy': lead['policy'],
            'pitch_notes': lead['pitch_notes'],
            'reasoning': lead['reasoning'],
            'profile': lead['profile'],
            'attempts': lead['attempts'],
            'attempts_remaining': min(
                config.MAX_CALL_ATTEMPTS - lead['attempts'],
                config.MAX_CALL_ATTEMPTS - lead['retry_rounds'],
            ),
        }

    first_calls, retries, exhausted, in_flight = [], [], [], []
    for lead in batch['leads'].values():
        if lead['stage'] == ANALYSED and lead['attempts'] == 0:
            first_calls.append(summary(lead))
        elif lead['stage'] == IN_PROGRESS:
            # Dialled and still live. Not ready, not retryable — handing this
            # lead back out is how somebody gets a second call while the first
            # one is ringing.
            in_flight.append({'lead_id': lead['lead_id'],
                              'call_id': lead['call_id']})
        elif lead['stage'] in (UNATTEMPTED, FAILED):
            if (lead['attempts'] < config.MAX_CALL_ATTEMPTS
                    and lead['retry_rounds'] < config.MAX_CALL_ATTEMPTS):
                retries.append(summary(lead))
            else:
                exhausted.append(lead['lead_id'])

    return {
        'status': 'success',
        'batch_id': batch['batch_id'],
        'first_calls': first_calls,
        'retries': retries,
        'retry_limit': config.MAX_CALL_ATTEMPTS,
        'exhausted': exhausted,
        'awaiting_result': in_flight,
        'counts': _counts(batch),
    }


def confirm_calling(tool_context: ToolContext) -> dict:
    """Unlock outbound calling for this batch, after the human has said yes.

    Only call this once a human has actually answered the confirmation
    question. It is not a formality — nothing recalls a voice call.
    """
    batch = _batch(tool_context)
    if batch is None:
        return _no_batch()
    if batch['confirmed']:
        return {'status': 'success', 'already_confirmed': True,
                'counts': _counts(batch)}
    batch['confirmed'] = True
    batch['confirmed_at'] = _now()
    _save(tool_context, batch)
    logger.info('%s confirmed for calling', batch['batch_id'])
    return {'status': 'success', 'counts': _counts(batch)}


def record_outreach_results(results: list[dict],
                            tool_context: ToolContext) -> dict:
    """Record what the Outreach Agent reported back for each call.

    Args:
        results: One dict per lead the specialist attempted. Each needs a
            `lead_id` and an `outcome` of 'contacted', 'unattempted',
            'failed' or 'in_progress'. `attempts` — how many times it actually
            dialled for this result — is optional but should be given whenever
            the specialist retried internally, or the retry cap will not bind.
            `detail` and `call_id` are optional.

    Returns:
        The counts after recording, and which leads are now owed a retry.
    """
    batch = _batch(tool_context)
    if batch is None:
        return _no_batch()

    recorded, unknown, unclear = [], [], []
    for result in results or []:
        if not isinstance(result, dict):
            unclear.append({'result': result, 'reason': 'not a record'})
            continue
        lead_id = str(result.get('lead_id') or '')
        lead = batch['leads'].get(lead_id)
        if lead is None:
            unknown.append(lead_id or '(missing lead_id)')
            continue

        outcome = (result.get('outcome') or '').strip().lower()
        detail = (result.get('detail') or '').strip()
        if result.get('call_id'):
            lead['call_id'] = str(result['call_id'])
        # Sticky: once any result for this lead was simulated, the lead's
        # history is a simulation and the report has to say so.
        if result.get('mock'):
            lead['mock'] = True

        # How many times the specialist actually dialled for this result. It
        # runs its own retry loop inside one invocation, so a single record can
        # stand for three calls — counting it as one is how the cap stops
        # binding and a lead gets dialled three times, handed back, and dialled
        # three times again.
        #
        # Taken verbatim when it is given, including zero: a poll that resolves
        # a call already placed reports zero, and charging it an attempt would
        # spend the budget on looking. Absent, it is the one call it reported.
        try:
            spent = max(0, int(result['attempts'])) if 'attempts' in result else 1
        except (TypeError, ValueError):
            spent = 1

        if outcome in ('contacted', 'answered', 'connected', 'success'):
            lead['stage'] = CALLED
            lead['attempts'] += spent
        elif outcome in ('in_progress', 'dispatched', 'queued', 'ringing',
                         'dialing'):
            # Live. Parked until someone polls it; not retryable meanwhile.
            lead['stage'] = IN_PROGRESS
            lead['attempts'] += spent
        elif outcome == 'unattempted':
            lead['stage'] = UNATTEMPTED
            lead['attempts'] += spent
            lead['retry_rounds'] += 1
        elif outcome in ('failed', 'no_answer', 'busy', 'voicemail'):
            lead['stage'] = FAILED
            lead['attempts'] += spent
            lead['retry_rounds'] += 1
        else:
            # An outcome nobody here recognises. Guessing it is what turns an
            # unreached lead into a reported contact — but leaving the lead
            # where it was is worse than it looks: at `analysed` with no
            # attempt spent it goes back out as a *first* call, and if the
            # specialist did dial it, the lead gets phoned twice. We do not
            # know what happened to this call, so nobody dials it again until
            # a human has looked.
            lead['stage'] = FLAGGED
            lead['last_status'] = outcome
            lead['note'] = (
                f'outreach reported an outcome we cannot read ({outcome!r}); '
                'unknown whether the call was placed'
            )
            unclear.append({'lead_id': lead_id, 'outcome': outcome})
            continue

        lead['last_status'] = outcome
        if detail:
            lead['note'] = detail

        # Two ceilings, because a specialist reporting zero attempts every
        # round — a platform that is down, say — would never reach the first
        # one, and the batch would circle until the conversation was abandoned.
        # `retry_rounds` bounds the loop by how many times we have come back to
        # this lead, whatever it says it spent.
        if lead['stage'] in (UNATTEMPTED, FAILED) and (
            lead['attempts'] >= config.MAX_CALL_ATTEMPTS
            or lead['retry_rounds'] >= config.MAX_CALL_ATTEMPTS
        ):
            lead['stage'] = FLAGGED
            lead['note'] = (
                f'{lead["attempts"]} attempts over {lead["retry_rounds"]} '
                f'rounds, last outcome {outcome}. Needs a human.'
            )
        recorded.append(lead_id)

    _save(tool_context, batch)
    counts = _counts(batch)
    return {
        'status': 'success',
        'recorded': len(recorded),
        'unknown_lead_ids': unknown,
        'unclear_outcomes': unclear,
        'counts': counts,
        'pending_retry': counts['pending_retry'],
    }


def flag_for_human_review(lead_id: str, reason: str,
                          tool_context: ToolContext) -> dict:
    """Take one lead out of the pipeline and leave it for a human.

    Use this when a sub-agent keeps erroring on a lead, when the data is
    unusable, or when retries are exhausted.
    """
    batch = _batch(tool_context)
    if batch is None:
        return _no_batch()
    lead = batch['leads'].get(str(lead_id))
    if lead is None:
        return {
            'status': 'error',
            'error_message': f'No lead {lead_id!r} in this batch.',
        }
    lead['stage'] = FLAGGED
    lead['note'] = reason
    _save(tool_context, batch)
    return {'status': 'success', 'lead_id': lead_id, 'counts': _counts(batch)}


def batch_status(tool_context: ToolContext) -> dict:
    """The state of the batch: counts, and every lead's stage.

    This is the only source for a figure you report to a human.
    """
    batch = _batch(tool_context)
    if batch is None:
        return _no_batch()
    return {
        'status': 'success',
        'batch_id': batch['batch_id'],
        'source': batch['source'],
        'opened_at': batch['opened_at'],
        'confirmed_for_calling': batch['confirmed'],
        'calls_are_simulated': config.mock_calls(),
        'counts': _counts(batch),
        'leads': [
            {
                'lead_id': lead['lead_id'],
                'name': lead['name'],
                'stage': lead['stage'],
                'priority': lead.get('priority', ''),
                'gap_k': lead.get('gap_k', 0),
                'age': lead.get('age', 0),
                'policy': lead['policy'],
                'pitch_notes': lead['pitch_notes'],
                'attempts': lead['attempts'],
                'call_id': lead['call_id'],
                'simulated': lead['mock'],
                'last_status': lead['last_status'],
                'note': lead['note'],
            }
            for lead in batch['leads'].values()
        ],
    }
